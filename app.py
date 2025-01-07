from datetime import datetime
from collections import defaultdict
import pytz
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from itsdangerous import URLSafeTimedSerializer
from flask_sqlalchemy import SQLAlchemy
from pytz import timezone, utc
from flask_mail import Mail, Message
import os
import csv
import io
from flask import Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from flask import send_file
from flask import make_response
import re

app = Flask(__name__)

# Configuration
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///training.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your_secret_key'

# Initialize extensions
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
migrate = Migrate(app, db)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Serializer for secure tokens
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])

# Initialize Flask-Mail
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 'training.tracker.noreply@gmail.com'  # Gmail account
app.config['MAIL_PASSWORD'] = 'enez gchs gfdw xdyh'  # Gmail app password
app.config['MAIL_DEFAULT_SENDER'] = 'training.tracker.noreply@gmail.com'

mail = Mail(app)


# Define database models
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    first_name = db.Column(db.String(50), nullable=False)  # First Name
    last_name = db.Column(db.String(50), nullable=False)  # Last Name
    email = db.Column(db.String(120), unique=True, nullable=False)
    phone_number = db.Column(db.String(15), nullable=True)
    password = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'admin' or 'trainer'
    is_active = db.Column(db.Boolean, default=False)


class TrainingAssignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    team_member_id = db.Column(db.Integer, db.ForeignKey('team_member.id'), nullable=False)
    trainer_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    position_id = db.Column(db.Integer, db.ForeignKey('position.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    timeframe = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    team_member = db.relationship('TeamMember', backref='assignments')
    trainer = db.relationship('User', backref='assignments')
    position = db.relationship('Position', backref='assignments')


class Comment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.Integer, db.ForeignKey('training_session.id'), nullable=False)
    trainer_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    parent_id = db.Column(db.Integer, db.ForeignKey('comment.id'), nullable=True)

    # Relationships
    trainer = db.relationship('User', backref='comments', lazy=True)  # Trainer/User relationship
    replies = db.relationship(
        'Comment',
        backref=db.backref('parent', remote_side=[id]),
        lazy='joined',  # Load replies with the parent comment
        cascade="all, delete-orphan"  # Handle cascading deletes for replies
    )


class TeamMember(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    start_date = db.Column(db.Date, nullable=False, default=datetime.utcnow)
    training_sessions = db.relationship('TrainingSession', backref='team_member', lazy=True)


class Position(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    training_sessions = db.relationship('TrainingSession', backref='position', lazy=True)


class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    message = db.Column(db.String(255), nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class TrainingSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    team_member_id = db.Column(db.Integer, db.ForeignKey('team_member.id'), nullable=False)
    position_id = db.Column(db.Integer, db.ForeignKey('position.id'), nullable=False)
    trainer_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    trainer = db.relationship('User', backref='training_sessions', lazy=True)


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# Utility functions for validation
def validate_username(username):
    pattern = r'^[a-zA-Z0-9~`!@#$%^&*()+=_\-{}[\]|:;”’?/<>,.]{6,20}$'
    return re.match(pattern, username) and " " not in username


def validate_password(password):
    pattern = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[~`!@#$%^&*()+=_\-{}[\]|:;”’?/<>,.]).{6,20}$'
    return re.match(pattern, password)


# Routes
@app.route('/')
def home():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and bcrypt.check_password_hash(user.password, password):
            login_user(user)
            flash(f'Welcome {user.first_name}!', 'success')

            # Redirect based on role
            if user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif user.role == 'trainer':
                return redirect(url_for('trainer_dashboard'))
            else:
                flash('Unknown role. Please contact support.', 'danger')
                return redirect(url_for('login'))

        flash('Invalid username or password.', 'danger')
    return render_template('login.html')


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email')
        user = User.query.filter_by(email=email).first()
        if user:
            token = serializer.dumps(email, salt='password-reset')
            reset_link = url_for('reset_password', token=token, _external=True)

            # Send email
            msg = Message(
                subject="Password Reset Request",
                recipients=[email],
                body=f"Click the link to reset your password: {reset_link}"
            )
            mail.send(msg)
            flash("Password reset link sent to your email.", "success")
        else:
            flash("No account found with that email address.", "danger")
    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        email = serializer.loads(token, salt='password-reset', max_age=3600)
    except Exception:
        flash("The reset link is invalid or has expired.", "danger")
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        if password != confirm_password:
            flash("Passwords do not match.", "danger")
            return redirect(url_for('reset_password', token=token))

        user = User.query.filter_by(email=email).first()
        if user:
            hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
            user.password = hashed_password
            db.session.commit()
            flash("Your password has been reset. You can now log in.", "success")
            return redirect(url_for('login'))
        else:
            flash("No account found for this email.", "danger")
            return redirect(url_for('forgot_password'))

    return render_template('reset_password.html', token=token)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))


@app.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    if current_user.role == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif current_user.role == 'trainer':
        return redirect(url_for('trainer_dashboard'))
    else:
        return "Unauthorized", 403


@app.route('/api/notifications/mark-read', methods=['POST'])
@login_required
def mark_notifications_as_read():
    notifications = Notification.query.filter_by(user_id=current_user.id, is_read=False).all()
    for notification in notifications:
        notification.is_read = True
    db.session.commit()
    return jsonify({"success": True})


@app.route('/trainer-dashboard', methods=['GET'])
@login_required
def trainer_dashboard():
    if current_user.role != 'trainer':
        return "Unauthorized", 403

    query = request.args.get('query', '').strip()
    sessions_count = TrainingSession.query.filter_by(trainer_id=current_user.id).count()
    positions = Position.query.all()

    # Filter team members based on search query
    if query:
        team_members = TeamMember.query.filter(TeamMember.name.ilike(f"%{query}%")).all()
    else:
        team_members = TeamMember.query.all()

    return render_template(
        'trainer_dashboard.html',
        sessions_count=sessions_count,
        team_members=team_members,
        positions=positions,
        query=query
    )


@app.route('/admin-dashboard', methods=['GET'])
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        return "Unauthorized", 403

    team_members_count = TeamMember.query.count()
    positions_count = Position.query.count()
    users_count = User.query.count()

    return render_template(
        'admin_dashboard.html',
        team_members_count=team_members_count,
        positions_count=positions_count,
        users_count=users_count
    )


@app.route('/api/assignments')
@login_required
def get_assignments():
    if current_user.role == 'admin':
        assignments = TrainingAssignment.query.all()
    elif current_user.role == 'trainer':
        assignments = TrainingAssignment.query.filter_by(trainer_id=current_user.id).all()
    else:
        return "Unauthorized", 403

    data = [
        {
            "id": assignment.id,
            "title": f"{assignment.team_member.name} - {assignment.position.name}",
            "start": assignment.date.isoformat(),
            "time": assignment.time.strftime('%H:%M'),
            "trainer": f"{assignment.trainer.first_name} {assignment.trainer.last_name}",
        }
        for assignment in assignments
    ]
    return jsonify(data)


@app.route('/api/notifications', methods=['GET'])
@login_required
def get_notifications():
    notifications = Notification.query.filter_by(user_id=current_user.id, is_read=False).all()
    return jsonify([{"id": n.id, "message": n.message, "created_at": n.created_at} for n in notifications])


def send_assignment_email(trainer_email, team_member_name, position_name, date, timeframe):
    message = Message(
        subject="Training Assignment Notification",
        recipients=[trainer_email],
        body=f"""
        Hello,

        You have been assigned to train {team_member_name} for the position of {position_name}.

        Date: {date.strftime('%Y-%m-%d')}
        Timeframe: {timeframe}

        Please ensure you're available for this session.

        Best regards,
        Admin Team
        """
    )
    mail.send(message)


@app.route('/trainer/calendar', methods=['GET'])
@login_required
def trainer_calendar():
    if current_user.role != 'trainer':
        return "Unauthorized", 403
    return render_template('trainer_calendar.html')


@app.route('/trainer/assignments-data', methods=['GET'])
@login_required
def trainer_assignments_data():
    if current_user.role != 'trainer':
        return "Unauthorized", 403

    assignments = TrainingAssignment.query.filter_by(trainer_id=current_user.id).all()

    events = []
    timeframe_map = {
        "breakfast": ("06:00", "10:30"),
        "lunch": ("11:00", "14:00"),
        "thru-period": ("14:00", "16:00"),
        "dinner": ("16:00", "19:00"),
        "close": ("19:00", "21:00"),
    }

    for assignment in assignments:
        start_time, end_time = timeframe_map[assignment.timeframe]
        start_datetime = datetime.combine(assignment.date, datetime.strptime(start_time, "%H:%M").time())
        end_datetime = datetime.combine(assignment.date, datetime.strptime(end_time, "%H:%M").time())
        events.append({
            'title': f"{assignment.team_member.name} ({assignment.position.name})",
            'start': start_datetime.isoformat(),
            'end': end_datetime.isoformat(),
            'extendedProps': {
                'team_member': assignment.team_member.name,
                'position': assignment.position.name,
                'timeframe': assignment.timeframe,
            },
        })

    return jsonify(events)


@app.route('/admin/unassign-session/<int:session_id>', methods=['POST'])
@login_required
def unassign_session(session_id):
    if current_user.role != 'admin':
        return "Unauthorized", 403

    # Fetch the session
    session = TrainingAssignment.query.get_or_404(session_id)

    # Explicitly load related objects before deleting the session
    team_member = session.team_member
    position = session.position
    trainer = session.trainer

    # Delete the session
    db.session.delete(session)
    db.session.commit()

    # Create a notification for the trainer
    notification_message = f"You have been unassigned from the training session with {team_member.name} for {position.name} on {session.date.strftime('%Y-%m-%d')}."
    notification = Notification(
        user_id=trainer.id,
        message=notification_message
    )
    db.session.add(notification)
    db.session.commit()

    # Send email notification to the trainer
    try:
        msg = Message(
            subject="You have been unassigned from a training session",
            recipients=[trainer.email],
            body=f"""
            Dear {trainer.first_name},

            You have been unassigned from the following training session:

            - Team Member: {team_member.name}
            - Position: {position.name}
            - Date: {session.date.strftime('%d-%m-%y')}

            Please reach out to the admin if you have any questions.

            Regards,
            Training Management System
            """
        )
        mail.send(msg)
    except Exception as e:
        app.logger.error(f"Error sending email notification: {str(e)}")

    flash('Training session unassigned successfully, and the trainer has been notified!',
          'success')
    return redirect(request.referrer or url_for('admin_users'))


@app.route('/admin/assignments-data', methods=['GET'])
@login_required
def fetch_assignments():
    if current_user.role != 'admin':
        return "Unauthorized", 403

    start = request.args.get('start')
    end = request.args.get('end')

    # Convert start and end to datetime
    start_date = datetime.fromisoformat(start[:-6])  # Remove timezone offset for parsing
    end_date = datetime.fromisoformat(end[:-6])

    # Map timeframe to specific hours
    timeframe_map = {
        'breakfast': ('06:00', '10:30'),
        'lunch': ('11:00', '14:00'),
        'thru-period': ('14:00', '16:00'),  # Add this missing key
        'dinner': ('16:00', '19:00'),
        'close': ('19:00', '21:00'),
    }

    assignments = TrainingAssignment.query.filter(
        TrainingAssignment.date >= start_date.date(),
        TrainingAssignment.date <= end_date.date()
    ).all()

    events = []
    for assignment in assignments:
        start_time, end_time = timeframe_map.get(assignment.timeframe,
                                                 ('00:00', '23:59'))  # Default for missing timeframe
        events.append({
            'title': f"{assignment.team_member.name} - {assignment.position.name}",
            'start': f"{assignment.date}T{start_time}:00",
            'end': f"{assignment.date}T{end_time}:00",
            'extendedProps': {
                'trainer': f"{assignment.trainer.first_name} {assignment.trainer.last_name}",
                'team_member': assignment.team_member.name,
                'position': assignment.position.name,
                'timeframe': assignment.timeframe,
            },
        })

    return jsonify(events)


@app.route('/admin/assign-trainer', methods=['GET', 'POST'])
@login_required
def assign_trainer():
    if current_user.role != 'admin':
        return "Unauthorized", 403

    if request.method == 'POST':
        team_member_id = request.form.get('team_member_id')
        trainer_id = request.form.get('trainer_id')
        position_id = request.form.get('position_id')
        date = request.form.get('date')
        timeframe = request.form.get('timeframe')

        if not (team_member_id and trainer_id and position_id and date and timeframe):
            flash("All fields are required.", "danger")
            return redirect(url_for('assign_trainer'))

        # Create a new assignment
        new_assignment = TrainingAssignment(
            team_member_id=team_member_id,
            trainer_id=trainer_id,
            position_id=position_id,
            date=datetime.strptime(date, '%Y-%m-%d').date(),
            timeframe=timeframe,
        )
        db.session.add(new_assignment)

        # Create a notification for the assigned trainer
        trainer = User.query.get(trainer_id)
        team_member = TeamMember.query.get(team_member_id)
        position = Position.query.get(position_id)
        notification_message = f"You have been assigned to train {team_member.name} as a {position.name} on {date} during {timeframe}."
        notification = Notification(
            user_id=trainer.id,
            message=notification_message
        )
        db.session.add(notification)

        # Commit all changes
        db.session.commit()

        # Send email notification (optional)
        send_assignment_email(
            trainer.email,
            team_member.name,
            position.name,
            new_assignment.date,
            timeframe
        )

        flash("Trainer assigned and notified successfully!", "success")
        return redirect(url_for('assign_trainer'))

    team_members = TeamMember.query.all()
    trainers = User.query.filter_by(role='trainer').all()
    positions = Position.query.all()

    return render_template(
        'assign_trainer.html',
        team_members=team_members,
        trainers=trainers,
        positions=positions
    )


@app.route('/admin/add-user', methods=['GET', 'POST'])
@login_required
def add_user():
    if current_user.role != 'admin':
        return "Unauthorized", 403

    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        phone_number = request.form['phone_number']
        password = request.form['password']
        role = request.form['role']

        # Validate username and password
        if not validate_username(username):
            flash('Invalid username. Follow the rules.', 'danger')
            return redirect(url_for('add_user'))
        if not validate_password(password):
            flash('Password must meet complexity requirements.', 'danger')
            return redirect(url_for('add_user'))

        # Check for existing user
        existing_user = User.query.filter((User.username == username) | (User.email == email)).first()
        if existing_user:
            flash('Username or email already exists. Please use a different one.', 'danger')
            return redirect(url_for('add_user'))

        # Hash the password
        hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')

        # Create new user
        new_user = User(
            username=username,
            email=email,
            phone_number=phone_number,
            password=hashed_password,
            role=role,
            is_active=True  # Immediate activation for manual addition
        )
        db.session.add(new_user)
        db.session.commit()

        flash('User added successfully!', 'success')
        return redirect(url_for('list_users'))

    return render_template('add_user.html')


@app.route('/search-team-members', methods=['GET', 'POST'])
@login_required
def search_team_members():
    if current_user.role not in ['admin', 'trainer']:
        return "Unauthorized", 403

    query = request.args.get('query', '')
    if query:
        team_members = TeamMember.query.filter(TeamMember.name.ilike(f'%{query}%')).all()
    else:
        team_members = TeamMember.query.all()

    return render_template('search_team_members.html', team_members=team_members, query=query)


@app.route('/add-comment/<int:session_id>', methods=['GET', 'POST'])
@login_required
def add_comment_prompt(session_id):
    if current_user.role != 'trainer':
        return "Unauthorized", 403

    session = TrainingSession.query.get_or_404(session_id)

    # Convert the session timestamp to EST
    est = timezone('US/Eastern')
    if session.timestamp.tzinfo is None:
        # If the timestamp is naive, assume it's in UTC and localize it
        session.timestamp = pytz.utc.localize(session.timestamp).astimezone(est)
    else:
        # If already timezone-aware, convert it to EST
        session.timestamp = session.timestamp.astimezone(est)

    # Debugging: print timestamps for verification
    print(f"Original Timestamp (UTC): {session.timestamp.astimezone(pytz.utc)}")
    print(f"Converted Timestamp (EST): {session.timestamp}")

    if request.method == 'POST':
        content = request.form.get('content')
        if not content:
            flash("Comment cannot be empty.", "danger")
            return redirect(url_for('add_comment_prompt', session_id=session_id))

        new_comment = Comment(
            session_id=session_id,
            trainer_id=current_user.id,
            content=content
        )
        db.session.add(new_comment)
        db.session.commit()
        flash("Comment added successfully!", "success")
        return redirect(url_for('trainer_dashboard'))

    return render_template('add_comment_prompt.html', session=session)


@app.route('/session/<int:session_id>/comment', methods=['GET', 'POST'])
@login_required
def add_comment(session_id):
    session = TrainingSession.query.get_or_404(session_id)

    # Ensure the session.timestamp is timezone-aware and convert it to EST
    est = timezone('US/Eastern')
    if session.timestamp.tzinfo is None:
        session.timestamp = pytz.utc.localize(session.timestamp).astimezone(est)
    else:
        session.timestamp = session.timestamp.astimezone(est)

    if request.method == 'POST':
        content = request.form['content']
        parent_id = request.form.get('parent_id', None)  # Default to None for top-level comments

        # Add new comment or reply
        new_comment = Comment(
            session_id=session_id,
            trainer_id=current_user.id,
            content=content,
            parent_id=int(parent_id) if parent_id else None  # Explicitly set None if no parent
        )
        db.session.add(new_comment)
        db.session.commit()
        flash('Comment added successfully!', 'success')
        return redirect(url_for('view_session_details', session_id=session_id))

    # Fetch only top-level comments and their nested replies
    comments = (
        Comment.query.filter_by(session_id=session_id, parent_id=None)  # Only top-level comments
        .options(
            db.joinedload(Comment.replies).joinedload(Comment.trainer),  # Load nested replies and their trainers
            db.joinedload(Comment.trainer),  # Load trainer for top-level comments
        )
        .order_by(Comment.timestamp.desc())
        .all()
    )

    return render_template(
        'session_details.html',
        session=session,
        comments=comments,
    )


@app.route('/session/<int:session_id>', methods=['GET'])
@login_required
def view_session_details(session_id):
    session = TrainingSession.query.get_or_404(session_id)

    # Convert the timestamp to EST
    est = timezone('US/Eastern')
    if session.timestamp.tzinfo is None:
        session.timestamp = pytz.utc.localize(session.timestamp).astimezone(est)
    else:
        session.timestamp = session.timestamp.astimezone(est)

    # Fetch only top-level comments and their replies
    comments = (
        Comment.query.filter_by(session_id=session_id, parent_id=None)  # Exclude replies
        .options(
            db.joinedload(Comment.replies).joinedload(Comment.trainer),
            db.joinedload(Comment.trainer)
        )
        .order_by(Comment.timestamp.desc())
        .all()
    )

    return render_template(
        'session_details.html',
        session=session,
        comments=comments,
    )


@app.route('/admin/user/<int:user_id>', methods=['GET'])
@login_required
def user_details(user_id):
    if current_user.role != 'admin':
        return "Unauthorized", 403

    user = User.query.get_or_404(user_id)
    assigned_sessions = TrainingAssignment.query.filter_by(trainer_id=user_id).all()

    return render_template('user_details.html', user=user, assigned_sessions=assigned_sessions)


@app.route('/admin/invite-user', methods=['GET', 'POST'])
@login_required
def invite_user():
    if current_user.role != 'admin':
        return "Unauthorized", 403

    if request.method == 'POST':
        email = request.form['email']
        phone_number = request.form['phone_number']
        role = request.form['role']

        # Generate invite token
        token = serializer.dumps(email, salt='invite-token')
        invite_link = url_for('register', token=token, _external=True)

        # Send email invite
        try:
            msg = Message(
                subject='You Are Invited to Join as a Trainer',
                recipients=[email],
                body=f"Hello,\n\nYou have been invited to join as a trainer. "
                     f"Please click the following link to complete your registration:\n\n{invite_link}\n\n"
                     f"If you did not expect this email, please ignore it.\n\n"
                     f"Best regards,\nYour Admin Team"
            )
            mail.send(msg)
            flash(f'Invite sent to {email}', 'success')
        except Exception as e:
            flash(f'Failed to send invite. Error: {str(e)}', 'danger')

        return redirect(url_for('dashboard'))

    return render_template('invite_user.html')


@app.template_filter('format_phone')
def format_phone(value):
    """Format a 10-digit phone number as +1 (123) 456-7890"""
    if not value or len(value) != 10 or not value.isdigit():
        return value  # Return unformatted if input is invalid
    return f"+1 ({value[:3]}) {value[3:6]}-{value[6:]}"


@app.route('/admin/edit-user/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    if current_user.role != 'admin':
        return "Unauthorized", 403

    user = User.query.get_or_404(user_id)

    # Update user details from the form
    user.first_name = request.form['first_name']
    user.last_name = request.form['last_name']
    user.email = request.form['email']
    user.phone_number = request.form.get('phone_number', None)
    user.role = request.form['role']

    # Commit changes
    db.session.commit()
    flash(f"User details for {user.first_name} {user.last_name} updated successfully!", "success")
    return redirect(url_for('user_details', user_id=user.id))


@app.route('/register/<token>', methods=['GET', 'POST'])
def register(token):
    try:
        email = serializer.loads(token, salt='invite-token', max_age=3600)
    except:
        flash('The invite link is invalid or has expired.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        first_name = request.form['first_name']
        last_name = request.form['last_name']
        username = request.form['username']
        password = request.form['password']

        # Validate username and password
        if not validate_username(username):
            flash('Invalid username. Follow the rules.', 'danger')
            return redirect(request.url)
        if not validate_password(password):
            flash('Password must meet complexity requirements.', 'danger')
            return redirect(request.url)

        hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
        new_user = User(
            first_name=first_name,
            last_name=last_name,
            username=username,
            email=email,
            password=hashed_password,
            role='trainer',
            is_active=True
        )
        db.session.add(new_user)
        db.session.commit()

        flash('Account created successfully! You can now log in.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')


@app.route('/admin/add-team-member', methods=['GET', 'POST'])
@login_required
def add_team_member():
    if current_user.role != 'admin':
        return "Unauthorized", 403
    if request.method == 'POST':
        name = request.form['name']
        start_date_str = request.form['start_date']
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        new_member = TeamMember(name=name, start_date=start_date)
        db.session.add(new_member)
        db.session.commit()
        flash('Team member added successfully!', 'success')
        return redirect(url_for('list_team_members'))
    return render_template('add_team_member.html')


@app.route('/admin/add-position', methods=['GET', 'POST'])
@login_required
def add_position():
    if current_user.role != 'admin':
        return "Unauthorized", 403
    if request.method == 'POST':
        name = request.form['name']
        new_position = Position(name=name)
        db.session.add(new_position)
        db.session.commit()
        flash('Position added successfully!', 'success')
        return redirect(url_for('list_positions'))
    return render_template('add_position.html')


@app.route('/admin/positions')
@login_required
def list_positions():
    if current_user.role != 'admin':
        return "Unauthorized", 403
    positions = Position.query.all()
    return render_template('list_positions.html', positions=positions)


@app.route('/admin/users', methods=['GET'])
@login_required
def list_users():
    if current_user.role != 'admin':
        return "Unauthorized", 403
    users = User.query.all()  # Fetch all users from the database
    return render_template('list_users.html', users=users)


@app.route('/admin/delete-team-member/<int:team_member_id>', methods=['POST'])
@login_required
def delete_team_member(team_member_id):
    if current_user.role != 'admin':
        return "Unauthorized", 403

    team_member = TeamMember.query.get_or_404(team_member_id)
    db.session.delete(team_member)
    db.session.commit()
    flash(f'Team member "{team_member.name}" deleted successfully.', 'success')
    return redirect(url_for('list_team_members'))


@app.route('/admin/delete-position/<int:position_id>', methods=['POST'])
@login_required
def delete_position(position_id):
    if current_user.role != 'admin':
        return "Unauthorized", 403
    position = Position.query.get_or_404(position_id)
    db.session.delete(position)
    db.session.commit()
    flash('Position deleted successfully.', 'success')
    return redirect(url_for('list_positions'))


@app.route('/admin/delete-user/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if current_user.role != 'admin':
        return "Unauthorized", 403

    user = User.query.get_or_404(user_id)

    # Prevent deleting the currently logged-in admin
    if user.id == current_user.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for('list_users'))

    db.session.delete(user)
    db.session.commit()

    flash(f"User '{user.first_name} {user.last_name}' deleted successfully.", 'success')
    return redirect(url_for('list_users'))


@app.route('/team-members')
@login_required
def list_team_members():
    team_members = TeamMember.query.all()
    return render_template('list_team_members.html', team_members=team_members)


@app.route('/team-member/<int:team_member_id>')
@login_required
def team_member_profile(team_member_id):
    if current_user.role not in ['admin', 'trainer']:
        return "Unauthorized", 403

    team_member = TeamMember.query.get_or_404(team_member_id)
    training_sessions = TrainingSession.query.filter_by(team_member_id=team_member_id).all()

    # Calculate training counts by position
    position_counts = defaultdict(int)
    for session in training_sessions:
        position_counts[session.position.name] += 1

    return render_template(
        'team_member_profile.html',
        team_member=team_member,
        training_sessions=training_sessions,
        position_counts=position_counts
    )


@app.route('/admin/reports', methods=['GET'])
@login_required
def reports():
    if current_user.role != 'admin':
        return "Unauthorized", 403

    # Fetch and process data
    sessions_per_member = [
        (row.name, row.count)
        for row in db.session.query(
            TeamMember.name, db.func.count(TrainingSession.id).label('count')
        )
        .join(TrainingSession, TeamMember.id == TrainingSession.team_member_id)
        .group_by(TeamMember.name)
        .order_by(db.func.count(TrainingSession.id).desc())
    ]

    sessions_per_position = [
        (row.name, row.count)
        for row in db.session.query(
            Position.name, db.func.count(TrainingSession.id).label('count')
        )
        .join(TrainingSession, Position.id == TrainingSession.position_id)
        .group_by(Position.name)
        .order_by(db.func.count(TrainingSession.id).desc())
    ]

    most_active_members = sessions_per_member[:5]  # Top 5 members

    # Ensure data is JSON serializable
    sessions_per_member = list(sessions_per_member)
    sessions_per_position = list(sessions_per_position)
    most_active_members = list(most_active_members)

    return render_template(
        'reports.html',
        sessions_per_member=sessions_per_member,
        sessions_per_position=sessions_per_position,
        most_active_members=most_active_members,
    )


@app.route('/admin/reports/export', methods=['GET'])
@login_required
def export_reports():
    if current_user.role != 'admin':
        return "Unauthorized", 403

    # Fetch data
    sessions_per_member = db.session.query(
        TeamMember.name, db.func.count(TrainingSession.id).label('count')
    ).join(TrainingSession, TeamMember.id == TrainingSession.team_member_id) \
        .group_by(TeamMember.name).order_by(db.func.count(TrainingSession.id).desc()).all()

    sessions_per_position = db.session.query(
        Position.name, db.func.count(TrainingSession.id).label('count')
    ).join(TrainingSession, Position.id == TrainingSession.position_id) \
        .group_by(Position.name).order_by(db.func.count(TrainingSession.id).desc()).all()

    # Create an Excel workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reports"

    # Write data headers
    sheet.append(["Report Type", "Name", "Count"])

    # Write sessions per member
    sheet.append(["Sessions Per Member"])
    for member, count in sessions_per_member:
        sheet.append(["", member, count])

    # Leave a blank row
    sheet.append([])

    # Write sessions per position
    sheet.append(["Sessions Per Position"])
    for position, count in sessions_per_position:
        sheet.append(["", position, count])

    # Autofit columns
    for col in sheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get column letter
        for cell in col:
            try:  # Necessary to handle empty cells
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2  # Add some padding
        sheet.column_dimensions[column].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Send the Excel file as a response
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="reports.xlsx"
    )


@app.route('/log-session/<int:team_member_id>', methods=['POST'])
@login_required
def log_session(team_member_id):
    if current_user.role != 'trainer':
        return "Unauthorized", 403

    position_id = request.form.get('position_id')
    if not position_id:
        flash("Please select a position.", "danger")
        return redirect(url_for('trainer_dashboard'))

    new_session = TrainingSession(
        team_member_id=team_member_id,
        position_id=position_id,
        trainer_id=current_user.id
    )
    db.session.add(new_session)
    db.session.commit()
    flash('Training session logged successfully!', 'success')
    return redirect(url_for('add_comment_prompt', session_id=new_session.id))


@app.route('/sessions')
@login_required
def view_sessions():
    est = timezone('US/Eastern')
    sort_by = request.args.get('sort', 'date')  # Default to sorting by date
    sort_direction = request.args.get('direction', 'asc')  # Default to ascending

    # Map sorting keys to model attributes
    sort_options = {
        'name': TeamMember.name,
        'position': Position.name,
        'trainer': User.first_name,
        'date': TrainingSession.timestamp
    }

    # Get the sorting column, default to date if invalid key
    sort_column = sort_options.get(sort_by, TrainingSession.timestamp)

    # Apply sorting direction
    if sort_direction == 'desc':
        sort_column = sort_column.desc()

    # Fetch and sort sessions
    sessions = db.session.query(TrainingSession).join(TeamMember).join(Position).join(User).order_by(sort_column).all()

    # Convert each timestamp to EST
    for session in sessions:
        session.timestamp = session.timestamp.replace(tzinfo=pytz.utc).astimezone(est)

    return render_template('view_sessions.html', sessions=sessions)


def to_est(dt):
    est = timezone('US/Eastern')
    return dt.replace(tzinfo=pytz.utc).astimezone(est)


@app.context_processor
def inject_timezone():
    est_timezone = timezone('US/Eastern')
    return {'est_timezone': est_timezone}


@app.route('/init-db')
def init_db():
    db.create_all()
    # Add default admin user with first_name and last_name
    if not User.query.filter_by(username='admin').first():
        hashed_password = bcrypt.generate_password_hash('admin123').decode('utf-8')
        admin_user = User(
            username='admin',
            first_name='Admin',
            last_name='User',
            email='admin@example.com',
            phone_number='1234567890',
            password=hashed_password,
            role='admin',
            is_active=True
        )
        db.session.add(admin_user)
        db.session.commit()
    return "Database initialized with default admin!"


@app.route('/debug')
def debug_time():
    utc_time = datetime.utcnow()
    est_time = utc_time.astimezone(pytz.timezone('US/Eastern'))
    return f"UTC: {utc_time}, EST: {est_time}"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", 5000)))
